import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from time import sleep

# Autor: Conceição Larissa Barbosa Ferreira
# Data de Criação: 28/07/2024
#
# Descrição:
# Este script é um bot para atualizar os valores das ações em uma planilha Excel.
# O script lê uma lista de ações, consulta o valor atual de cada ação no Google,
# e atualiza a planilha com a data e o valor correspondente.
# O objetivo é manter um histórico diário dos valores das ações.
# Função para obter o valor da ação


def obter_valor_acao(driver, acao):
    driver.get('https://www.google.com.br/?hl=pt-BR')
    try:
        search_box = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, 'q')))
        search_box.send_keys(acao)
        sleep(2)
        search_box.send_keys(Keys.ENTER)
        price_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//span[@class="IsqQVc NprOob wT3VGc"]'))
        )
        return price_element.text
    except Exception as e:
        print(f"Erro ao obter o valor da ação {acao}: {e}")
        return None

# Configurações do WebDriver
options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Carrega a planilha e a data atual
caminho_planilha = 'C:\\Estudos\\Aprendizado_de_Python\\Busca_Acoes.xlsx'
workbook = openpyxl.load_workbook(caminho_planilha)
sheet = workbook.active
data_hoje = datetime.today().strftime('%d/%m/%Y')

# Preenche a data na próxima coluna disponível
ultima_coluna = sheet.max_column
sheet.cell(row=1, column=ultima_coluna + 1, value=data_hoje)

# Itera sobre as linhas da planilha para atualizar os valores das ações
for linha in range(2, sheet.max_row + 1):
    acao = sheet.cell(row=linha, column=1).value
    if acao:
        valor_acao = obter_valor_acao(driver, acao)
        if valor_acao:
            sheet.cell(row=linha, column=ultima_coluna + 1, value=valor_acao)

# Salva a planilha e fecha o navegador
workbook.save(caminho_planilha)
driver.quit()